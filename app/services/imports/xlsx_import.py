from __future__ import annotations

from collections import Counter
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import ImportBatchStatus, KeyStatus, Plan, VPNKey
from app.repositories.audit_logs import AuditLogsRepository
from app.repositories.import_batches import ImportBatchesRepository
from app.services.security import KeyProtector
from app.services.transactions import transactional


class XlsxImportService:
    required_columns = {"plan_code", "key_value"}
    optional_columns = {"external_ref", "comment", "expires_at"}

    def __init__(
        self,
        *,
        session: AsyncSession,
        import_batches_repo: ImportBatchesRepository,
        audit_logs_repo: AuditLogsRepository,
        key_protector: KeyProtector,
    ) -> None:
        self._session = session
        self._import_batches_repo = import_batches_repo
        self._audit_logs_repo = audit_logs_repo
        self._key_protector = key_protector

    async def preview(self, *, filename: str, content: bytes) -> dict:
        parsed = await self._parse_and_validate(content)
        return {
            "filename": filename,
            "rows_total": parsed["rows_total"],
            "rows_valid": len(parsed["valid_rows"]),
            "rows_rejected": len(parsed["errors"]),
            "errors": parsed["errors"],
        }

    async def import_file(self, *, filename: str, content: bytes, uploaded_by_user_id: int | None) -> dict:
        parsed = await self._parse_and_validate(content)
        async with transactional(self._session):
            batch = await self._import_batches_repo.create(
                uploaded_by_user_id=uploaded_by_user_id,
                source_filename=filename,
                rows_total=parsed["rows_total"],
                rows_imported=len(parsed["valid_rows"]),
                rows_rejected=len(parsed["errors"]),
                status=ImportBatchStatus.COMPLETED.value,
                report_json={"errors": parsed["errors"]},
            )
            for row in parsed["valid_rows"]:
                self._session.add(
                    VPNKey(
                        plan_id=row["plan_id"],
                        key_value_encrypted=self._key_protector.encrypt(row["key_value"]),
                        key_fingerprint=row["fingerprint"],
                        external_ref=row.get("external_ref"),
                        comment=row.get("comment"),
                        expires_at=row.get("expires_at"),
                        status=KeyStatus.AVAILABLE.value,
                        imported_batch_id=batch.id,
                    )
                )
            await self._audit_logs_repo.add(
                actor_user_id=uploaded_by_user_id,
                entity_type="import_batch",
                entity_id=str(batch.id),
                action="xlsx_import_completed",
                payload_json={
                    "batch_id": batch.id,
                    "filename": filename,
                    "rows_total": parsed["rows_total"],
                    "rows_imported": len(parsed["valid_rows"]),
                    "rows_rejected": len(parsed["errors"]),
                },
                correlation_id=f"import-batch:{batch.id}",
            )
        return {
            "batch_id": batch.id,
            "rows_total": parsed["rows_total"],
            "rows_imported": len(parsed["valid_rows"]),
            "rows_rejected": len(parsed["errors"]),
            "errors": parsed["errors"],
        }

    async def _parse_and_validate(self, content: bytes) -> dict:
        workbook = load_workbook(BytesIO(content))
        if "keys" not in workbook.sheetnames:
            raise ValueError("Workbook must contain a sheet named 'keys'")
        sheet = workbook["keys"]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Workbook is empty")

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        header_map = {header: index for index, header in enumerate(headers) if header}
        missing = self.required_columns - set(header_map)
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

        parsed_rows = []
        for row_index, raw_row in enumerate(rows[1:], start=2):
            plan_code = self._get_cell(raw_row, header_map, "plan_code")
            key_value = self._get_cell(raw_row, header_map, "key_value")
            external_ref = self._get_cell(raw_row, header_map, "external_ref")
            comment = self._get_cell(raw_row, header_map, "comment")
            expires_at = self._get_cell(raw_row, header_map, "expires_at")
            if not any([plan_code, key_value, external_ref, comment, expires_at]):
                continue
            parsed_rows.append(
                {
                    "row_number": row_index,
                    "plan_code": plan_code,
                    "key_value": key_value,
                    "external_ref": external_ref,
                    "comment": comment,
                    "expires_at": expires_at if isinstance(expires_at, datetime) else None,
                }
            )

        valid_rows: list[dict] = []
        errors: list[dict] = []
        plan_codes = {row["plan_code"] for row in parsed_rows if row["plan_code"]}
        plans = {
            plan.code: plan
            for plan in (
                await self._session.scalars(select(Plan).where(Plan.code.in_(plan_codes)))
            )
        }

        fingerprints = [
            self._key_protector.fingerprint(row["key_value"])
            for row in parsed_rows
            if row.get("key_value")
        ]
        fingerprint_counter = Counter(fingerprints)
        existing_fingerprints = set(
            await self._session.scalars(select(VPNKey.key_fingerprint).where(VPNKey.key_fingerprint.in_(fingerprints)))
        )

        for row in parsed_rows:
            row_errors: list[str] = []
            plan = plans.get(row["plan_code"])
            key_value = (row["key_value"] or "").strip()
            fingerprint = self._key_protector.fingerprint(key_value) if key_value else ""

            if not row["plan_code"]:
                row_errors.append("plan_code is required")
            elif plan is None:
                row_errors.append("plan_code does not exist")

            if not key_value:
                row_errors.append("key_value is required")
            elif fingerprint_counter[fingerprint] > 1:
                row_errors.append("duplicate key inside file")
            elif fingerprint in existing_fingerprints:
                row_errors.append("duplicate key already exists in DB")

            if row_errors:
                errors.append({"row_number": row["row_number"], "errors": row_errors})
                continue

            valid_rows.append(
                {
                    "row_number": row["row_number"],
                    "plan_id": plan.id,
                    "plan_code": plan.code,
                    "key_value": key_value,
                    "fingerprint": fingerprint,
                    "external_ref": (row["external_ref"] or "").strip() or None,
                    "comment": (row["comment"] or "").strip() or None,
                    "expires_at": row["expires_at"],
                }
            )

        return {
            "rows_total": len(parsed_rows),
            "valid_rows": valid_rows,
            "errors": errors,
        }

    @staticmethod
    def _get_cell(raw_row, header_map: dict[str, int], column_name: str):
        index = header_map.get(column_name)
        if index is None or index >= len(raw_row):
            return None
        value = raw_row[index]
        if isinstance(value, str):
            return value.strip()
        return value

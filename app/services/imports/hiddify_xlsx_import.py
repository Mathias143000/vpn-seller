from __future__ import annotations

from collections import Counter
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import HiddifyServer
from app.repositories.audit_logs import AuditLogsRepository
from app.services.hiddify import HiddifyService
from app.services.transactions import transactional


class HiddifyXlsxImportService:
    required_columns = {
        "name",
        "country_name",
        "base_url",
        "admin_proxy_path",
        "client_proxy_path",
        "api_key",
    }
    optional_columns = {"is_active"}

    def __init__(
        self,
        *,
        session: AsyncSession,
        hiddify: HiddifyService,
        audit_logs_repo: AuditLogsRepository,
    ) -> None:
        self._session = session
        self._hiddify = hiddify
        self._audit_logs_repo = audit_logs_repo

    async def preview(self, *, filename: str, content: bytes) -> dict:
        parsed = await self._parse_and_validate(content)
        return {
            "filename": filename,
            "rows_total": parsed["rows_total"],
            "rows_valid": len(parsed["valid_rows"]),
            "rows_rejected": len(parsed["errors"]),
            "errors": parsed["errors"],
        }

    async def import_file(self, *, filename: str, content: bytes, uploaded_by_user_id: int | None) -> dict:
        parsed = await self._parse_and_validate(content)
        imported_servers: list[dict] = []
        errors = list(parsed["errors"])

        for row in parsed["valid_rows"]:
            try:
                server = await self._hiddify.register_server(
                    name=row["name"],
                    country_name=row["country_name"],
                    base_url=row["base_url"],
                    admin_proxy_path=row["admin_proxy_path"],
                    client_proxy_path=row["client_proxy_path"],
                    api_key=row["api_key"],
                    actor_user_id=uploaded_by_user_id,
                    is_active=row["is_active"],
                )
                imported_servers.append(
                    {
                        "row_number": row["row_number"],
                        "server_id": server.id,
                        "name": server.name,
                        "country_name": server.country_name,
                    }
                )
            except Exception as exc:
                errors.append({"row_number": row["row_number"], "errors": [str(exc)]})

        async with transactional(self._session):
            await self._audit_logs_repo.add(
                actor_user_id=uploaded_by_user_id,
                entity_type="hiddify_import",
                entity_id=filename,
                action="hiddify_xlsx_import_completed",
                payload_json={
                    "filename": filename,
                    "rows_total": parsed["rows_total"],
                    "rows_imported": len(imported_servers),
                    "rows_rejected": len(errors),
                    "servers": imported_servers,
                },
            )

        return {
            "rows_total": parsed["rows_total"],
            "rows_imported": len(imported_servers),
            "rows_rejected": len(errors),
            "errors": errors,
            "servers": imported_servers,
        }

    async def _parse_and_validate(self, content: bytes) -> dict:
        workbook = load_workbook(BytesIO(content))
        if "servers" not in workbook.sheetnames:
            raise ValueError("Workbook must contain a sheet named 'servers'")
        sheet = workbook["servers"]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Workbook is empty")

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        header_map = {header: index for index, header in enumerate(headers) if header}
        missing = self.required_columns - set(header_map)
        if missing:
            raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

        parsed_rows = []
        for row_index, raw_row in enumerate(rows[1:], start=2):
            row_data = {
                "row_number": row_index,
                "name": self._get_cell(raw_row, header_map, "name"),
                "country_name": self._get_cell(raw_row, header_map, "country_name"),
                "base_url": self._get_cell(raw_row, header_map, "base_url"),
                "admin_proxy_path": self._get_cell(raw_row, header_map, "admin_proxy_path"),
                "client_proxy_path": self._get_cell(raw_row, header_map, "client_proxy_path"),
                "api_key": self._get_cell(raw_row, header_map, "api_key"),
                "is_active": self._get_cell(raw_row, header_map, "is_active"),
            }
            if not any(row_data[column] for column in row_data if column != "row_number"):
                continue
            parsed_rows.append(row_data)

        normalized_pairs = [
            (
                self._normalize_base_url(row["base_url"]),
                self._normalize_path(row["admin_proxy_path"]),
            )
            for row in parsed_rows
            if row.get("base_url") and row.get("admin_proxy_path")
        ]
        pair_counter = Counter(normalized_pairs)
        existing_pairs = {
            (server.base_url, server.admin_proxy_path)
            for server in (
                await self._session.scalars(select(HiddifyServer))
            )
        }

        valid_rows: list[dict] = []
        errors: list[dict] = []
        for row in parsed_rows:
            row_errors: list[str] = []
            base_url = self._normalize_base_url(row["base_url"])
            admin_proxy_path = self._normalize_path(row["admin_proxy_path"])
            client_proxy_path = self._normalize_path(row["client_proxy_path"])
            pair = (base_url, admin_proxy_path)

            if not (row["name"] or "").strip():
                row_errors.append("name is required")
            if not base_url:
                row_errors.append("base_url is required")
            if not admin_proxy_path:
                row_errors.append("admin_proxy_path is required")
            if not client_proxy_path:
                row_errors.append("client_proxy_path is required")
            if not (row["api_key"] or "").strip():
                row_errors.append("api_key is required")
            if pair != ("", "") and pair_counter[pair] > 1:
                row_errors.append("duplicate server inside file")
            elif pair in existing_pairs:
                row_errors.append("server already exists in DB")

            try:
                is_active = self._parse_bool(row["is_active"])
            except ValueError as exc:
                row_errors.append(str(exc))
                is_active = True

            if row_errors:
                errors.append({"row_number": row["row_number"], "errors": row_errors})
                continue

            valid_rows.append(
                {
                    "row_number": row["row_number"],
                    "name": row["name"].strip(),
                    "country_name": (row["country_name"] or "").strip() or "Без страны",
                    "base_url": base_url,
                    "admin_proxy_path": admin_proxy_path,
                    "client_proxy_path": client_proxy_path,
                    "api_key": row["api_key"].strip(),
                    "is_active": is_active,
                }
            )

        return {
            "rows_total": len(parsed_rows),
            "valid_rows": valid_rows,
            "errors": errors,
        }

    @staticmethod
    def _get_cell(raw_row, header_map: dict[str, int], column_name: str):
        index = header_map.get(column_name)
        if index is None or index >= len(raw_row):
            return None
        value = raw_row[index]
        if isinstance(value, str):
            return value.strip()
        return value

    @staticmethod
    def _normalize_base_url(value: str | None) -> str:
        return (value or "").strip().rstrip("/")

    @staticmethod
    def _normalize_path(value: str | None) -> str:
        return (value or "").strip().strip("/")

    @staticmethod
    def _parse_bool(value) -> bool:
        if value is None or value == "":
            return True
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().lower()
        if normalized in {"1", "true", "yes", "y", "on"}:
            return True
        if normalized in {"0", "false", "no", "n", "off"}:
            return False
        raise ValueError("is_active must be TRUE/FALSE")

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select

from app.db.models import AuditLog, Order, OrderStatus
from tests.conftest import build_services, create_available_key, create_user, make_xlsx, seed_default_plan


async def test_xlsx_import_validation_rejects_duplicate_and_unknown_plan(db, settings, fake_bot):
    plan = await seed_default_plan(db, settings)
    services = build_services(db, settings, fake_bot)
    await create_available_key(db, services["protector"], plan.id, "vpn://existing")
    content = make_xlsx(
        [
            {"plan_code": "plan_30", "key_value": "vpn://dup"},
            {"plan_code": "plan_30", "key_value": "vpn://dup"},
            {"plan_code": "unknown_plan", "key_value": "vpn://x"},
            {"plan_code": "plan_30", "key_value": "vpn://existing"},
        ]
    )

    preview = await services["xlsx_import"].preview(filename="keys.xlsx", content=content)

    assert preview["rows_valid"] == 0
    assert preview["rows_rejected"] == 4


async def test_xlsx_import_writes_audit_log(db, settings, fake_bot):
    await seed_default_plan(db, settings)
    user = await create_user(db, telegram_user_id=888001, role="admin")
    services = build_services(db, settings, fake_bot)
    content = make_xlsx([{"plan_code": "plan_30", "key_value": "vpn://audit-import"}])

    result = await services["xlsx_import"].import_file(
        filename="audit.xlsx",
        content=content,
        uploaded_by_user_id=user.id,
    )

    audit_logs = list((await db.scalars(select(AuditLog).order_by(AuditLog.id.asc()))))

    assert result["rows_imported"] == 1
    assert any(
        log.action == "xlsx_import_completed"
        and log.entity_type == "import_batch"
        and log.actor_user_id == user.id
        for log in audit_logs
    )


async def test_export_generation_contains_inventory_and_orders_with_audit(db, settings, fake_bot):
    plan = await seed_default_plan(db, settings)
    user = await create_user(db)
    services = build_services(db, settings, fake_bot)
    vpn_key = await create_available_key(db, services["protector"], plan.id, "vpn://exported")
    vpn_key.status = "issued"
    vpn_key.issued_to_user_id = user.id
    vpn_key.issued_at = datetime.now(tz=timezone.utc)
    order = Order(
        user_id=user.id,
        plan_id=plan.id,
        status=OrderStatus.ISSUED.value,
        amount_value=plan.price_value,
        amount_currency=plan.price_currency,
        payment_provider="fake",
        issued_key_id=vpn_key.id,
    )
    db.add(order)
    await db.commit()

    inventory_content = await services["xlsx_export"].export_inventory(status="issued", actor_user_id=user.id)
    inventory_workbook = load_workbook(BytesIO(inventory_content))
    inventory_sheet = inventory_workbook["keys"]
    inventory_rows = list(inventory_sheet.iter_rows(values_only=True))

    orders_content = await services["xlsx_export"].export_orders(actor_user_id=user.id)
    orders_workbook = load_workbook(BytesIO(orders_content))
    orders_sheet = orders_workbook["orders"]
    order_rows = list(orders_sheet.iter_rows(values_only=True))

    audit_logs = list((await db.scalars(select(AuditLog).order_by(AuditLog.id.asc()))))

    assert inventory_rows[1][1] == "plan_30"
    assert inventory_rows[1][3] == "vpn://exported"
    assert inventory_rows[0][-1] == "key_type"
    assert inventory_rows[1][-1] == "unknown"
    assert order_rows[1][0] == order.id
    assert order_rows[1][3] == "plan_30"
    assert order_rows[1][4] == OrderStatus.ISSUED.value
    assert any(log.action == "xlsx_inventory_exported" for log in audit_logs)
    assert any(log.action == "xlsx_orders_exported" for log in audit_logs)
